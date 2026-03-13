"""
Report generation — presets, date ranges, AADE data fetch, and Excel output.
Adapted from AgelClaw's accounting_xlsx.py for the Timologia multi-user app.
"""
import json
from collections import defaultdict
from datetime import datetime, date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import FERNET, REPORTS_DIR
from db import get_db


# ── Report presets ──

PRESETS = {
    "daily_summary": {
        "label": "Ημερήσια Σύνοψη",
        "description": "Έσοδα και έξοδα σήμερα",
    },
    "monthly_vat": {
        "label": "Μηνιαία ΦΠΑ",
        "description": "Ανάλυση ΦΠΑ τρέχοντος μήνα",
    },
    "quarterly_income": {
        "label": "Τριμηνιαία Έσοδα",
        "description": "Σύνοψη εσόδων τρέχοντος τριμήνου",
    },
    "annual_overview": {
        "label": "Ετήσια Επισκόπηση",
        "description": "Συνολική εικόνα τρέχοντος έτους",
    },
    "expenses_by_supplier": {
        "label": "Έξοδα ανά Προμηθευτή",
        "description": "Ανάλυση εξόδων ομαδοποιημένα ανά προμηθευτή",
    },
    "custom": {
        "label": "Προσαρμοσμένη",
        "description": "Επιλέξτε ημερομηνίες",
    },
}


def get_preset_dates(preset: str, params: dict) -> tuple[str, str]:
    """Return (date_from, date_to) based on preset name.

    For 'custom' preset, reads date_from/date_to from params dict.
    """
    today = date.today()

    if preset == "daily_summary":
        d = today.isoformat()
        return d, d

    elif preset == "monthly_vat":
        first = today.replace(day=1)
        return first.isoformat(), today.isoformat()

    elif preset == "quarterly_income":
        quarter_month = ((today.month - 1) // 3) * 3 + 1
        first = today.replace(month=quarter_month, day=1)
        return first.isoformat(), today.isoformat()

    elif preset == "annual_overview":
        first = today.replace(month=1, day=1)
        return first.isoformat(), today.isoformat()

    elif preset == "expenses_by_supplier":
        # Support relative periods for scheduled reports
        period = params.get("period")
        if period:
            return _resolve_relative_period(period, today)
        df = params.get("date_from")
        dt = params.get("date_to")
        if df and dt:
            return df, dt
        first = today.replace(month=1, day=1)
        return first.isoformat(), today.isoformat()

    elif preset == "custom":
        # Support relative periods for scheduled reports
        period = params.get("period")
        if period:
            return _resolve_relative_period(period, today)
        df = params.get("date_from")
        dt = params.get("date_to")
        if df and dt:
            return df, dt
        # Fallback: current year (for schedules created without params)
        first = today.replace(month=1, day=1)
        return first.isoformat(), today.isoformat()

    else:
        raise ValueError(f"Unknown preset: {preset}")


def _resolve_relative_period(period: str, today: date) -> tuple[str, str]:
    """Resolve relative period name to (date_from, date_to)."""
    import calendar

    if period == "today":
        d = today.isoformat()
        return d, d
    elif period == "yesterday":
        d = (today - timedelta(days=1)).isoformat()
        return d, d
    elif period == "last_7_days":
        return (today - timedelta(days=6)).isoformat(), today.isoformat()
    elif period == "last_30_days":
        return (today - timedelta(days=29)).isoformat(), today.isoformat()
    elif period == "current_month":
        return today.replace(day=1).isoformat(), today.isoformat()
    elif period == "previous_month":
        last_day_prev = today.replace(day=1) - timedelta(days=1)
        first_day_prev = last_day_prev.replace(day=1)
        return first_day_prev.isoformat(), last_day_prev.isoformat()
    elif period == "current_quarter":
        q_month = ((today.month - 1) // 3) * 3 + 1
        return today.replace(month=q_month, day=1).isoformat(), today.isoformat()
    elif period == "current_year":
        return today.replace(month=1, day=1).isoformat(), today.isoformat()
    else:
        raise ValueError(f"Unknown period: {period}")


async def generate_report(company_id: int, user_id: int, preset: str, params: dict) -> dict:
    """Fetch AADE data, generate xlsx, log to report_history, return result dict."""
    date_from, date_to = get_preset_dates(preset, params)

    # Load company + decrypt credentials
    with get_db() as conn:
        company = conn.execute("SELECT * FROM companies WHERE id=?", (company_id,)).fetchone()
    if not company:
        raise ValueError("Company not found")

    aade_uid = FERNET.decrypt(company["aade_user_id"].encode()).decode() if company["aade_user_id"] else ""
    aade_key = FERNET.decrypt(company["aade_subscription_key"].encode()).decode() if company["aade_subscription_key"] else ""
    if not aade_uid or not aade_key:
        raise ValueError("Company has no AADE credentials configured")

    from aade_client import MyDataClient
    client = MyDataClient(aade_uid, aade_key, company["afm"], env=company["aade_env"])

    # Direction filter: "all" (default), "sent", "received"
    direction_filter = params.get("direction", "all")

    try:
        income = []
        expenses = []
        if direction_filter in ("all", "sent"):
            income = await client.get_sent_invoices(date_from=date_from, date_to=date_to)
        if direction_filter in ("all", "received"):
            expenses = await client.get_received_invoices(date_from=date_from, date_to=date_to)
    finally:
        await client.close()

    # Filter out raw/error entries (those without a mark)
    income = [i for i in income if i.get("mark")]
    expenses = [e for e in expenses if e.get("mark")]

    # Resolve counterpart names (supplier names) from AFM via GSIS
    if income:
        _resolve_names(income, "sent")
    if expenses:
        _resolve_names(expenses, "received")

    # Generate Excel
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"report_{company['afm']}_{preset}_{timestamp}.xlsx"
    file_path = REPORTS_DIR / filename
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    if preset == "expenses_by_supplier":
        _generate_supplier_xlsx(file_path, expenses, company["afm"], date_from, date_to)
    else:
        _generate_xlsx(file_path, income, expenses, company["afm"], date_from, date_to)

    # Log to report_history
    with get_db() as conn:
        cur = conn.execute(
            """INSERT INTO report_history (company_id, user_id, preset, params, file_path, status)
               VALUES (?, ?, ?, ?, ?, 'success')""",
            (company_id, user_id, preset, json.dumps(params, ensure_ascii=False), str(file_path)),
        )
        report_id = cur.lastrowid

    return {
        "id": report_id,
        "preset": preset,
        "date_from": date_from,
        "date_to": date_to,
        "file_path": str(file_path),
        "filename": filename,
        "income_count": len(income),
        "expense_count": len(expenses),
    }


# ══════════════════════════════════════════════════════════════
# Excel generation — adapted from AgelClaw accounting_xlsx.py
# ══════════════════════════════════════════════════════════════

# ── Invoice type names ──

INVOICE_TYPE_NAMES = {
    "1.1": "Τιμολόγιο Πώλησης",
    "1.2": "Τιμολόγιο Πώλησης / Ενδ.",
    "1.3": "Τιμολόγιο Πώλησης / Τρίτων",
    "1.6": "Τιμολόγιο Αυτοπαράδοσης",
    "2.1": "ΤΠΥ",
    "2.2": "ΤΠΥ / Ενδ.",
    "2.3": "ΤΠΥ / Τρίτων",
    "2.4": "Συμβόλαιο - Έσοδο",
    "3.1": "Τίτλος Κτήσης",
    "5.1": "Πιστωτικό (συσχ.)",
    "5.2": "Πιστωτικό (μη συσχ.)",
    "11.1": "ΑΛΠ",
    "11.2": "ΑΠΥ",
}

# ── Styles ──

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
SUMMARY_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUMMARY_HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=11)
PROFIT_FONT = Font(bold=True, size=12, color="006100")
LOSS_FONT = Font(bold=True, size=12, color="9C0006")
LABEL_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
CURRENCY_FMT = '#,##0.00 €'


# ── Column definitions ──

INCOME_COLUMNS = [
    ("Ημ/νία", 12),
    ("Σειρά/ΑΑ", 14),
    ("Τύπος", 8),
    ("Περιγραφή", 28),
    ("ΑΦΜ Πελάτη", 14),
    ("Καθαρή Αξία", 15),
    ("ΦΠΑ", 12),
    ("Μικτή Αξία", 15),
    ("MARK", 14),
]

EXPENSE_COLUMNS = [
    ("Ημ/νία", 12),
    ("Σειρά/ΑΑ", 14),
    ("Τύπος", 8),
    ("Περιγραφή", 28),
    ("ΑΦΜ Προμηθευτή", 16),
    ("Επωνυμία Προμηθευτή", 32),
    ("Καθαρή Αξία", 15),
    ("ΦΠΑ", 12),
    ("Μικτή Αξία", 15),
    ("MARK", 14),
]


def _generate_xlsx(path: Path, income: list, expenses: list,
                   afm: str, date_from: str, date_to: str) -> Path:
    """Generate accounting Excel workbook with 3 sheets.

    Args:
        path: Output file path
        income: List of income invoice dicts (from AADE)
        expenses: List of expense invoice dicts (from AADE)
        afm: Business AFM
        date_from: Period start (YYYY-MM-DD)
        date_to: Period end (YYYY-MM-DD)

    Returns:
        Path to the generated file
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # Sheet 1: Έσοδα
    ws_income = wb.active
    ws_income.title = "Έσοδα"
    _write_invoice_sheet(ws_income, income, "income")

    # Sheet 2: Έξοδα
    ws_expenses = wb.create_sheet("Έξοδα")
    _write_invoice_sheet(ws_expenses, expenses, "expenses")

    # Sheet 3: Σύνοψη
    ws_summary = wb.create_sheet("Σύνοψη")
    _write_summary_sheet(ws_summary, income, expenses, afm, date_from, date_to)

    wb.save(str(path))
    return path


def _write_invoice_sheet(ws, invoices: list, direction: str):
    """Write income or expenses invoice sheet."""
    columns = INCOME_COLUMNS if direction == "income" else EXPENSE_COLUMNS

    # Column widths
    for i, (_, width) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Header row
    for col_idx, (title, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Data rows
    for row_idx, inv in enumerate(invoices, 2):
        net = _to_float(inv.get("totalNetValue", 0))
        vat = _to_float(inv.get("totalVatAmount", 0))
        gross = _to_float(inv.get("totalGrossValue", 0))
        inv_type = inv.get("invoiceType", "")
        type_name = INVOICE_TYPE_NAMES.get(inv_type, inv.get("type_name", inv_type))
        series_aa = f"{inv.get('series', '')}/{inv.get('aa', '')}"

        if direction == "expenses":
            values = [
                inv.get("issueDate", ""),
                series_aa,
                inv_type,
                type_name,
                inv.get("issuer_vat", ""),
                inv.get("issuer_name", ""),
                net,
                vat,
                gross,
                inv.get("mark", ""),
            ]
        else:
            values = [
                inv.get("issueDate", ""),
                series_aa,
                inv_type,
                type_name,
                inv.get("counterpart_vat", inv.get("issuer_vat", "")),
                net,
                vat,
                gross,
                inv.get("mark", ""),
            ]

        # Currency columns: net, vat, gross (3rd, 2nd, 1st before MARK)
        num_cols = len(columns)
        currency_cols = (num_cols - 3, num_cols - 2, num_cols - 1)

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            if col_idx in currency_cols:
                cell.number_format = CURRENCY_FMT
                cell.alignment = Alignment(horizontal="right")
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

    # Totals row
    if invoices:
        total_row = len(invoices) + 2
        num_cols = len(columns)
        ws.cell(row=total_row, column=4, value="ΣΥΝΟΛΟ").font = TOTAL_FONT

        for col_idx in (num_cols - 3, num_cols - 2, num_cols - 1):
            col_letter = get_column_letter(col_idx)
            cell = ws.cell(
                row=total_row, column=col_idx,
                value=f"=SUM({col_letter}2:{col_letter}{total_row - 1})",
            )
            cell.number_format = CURRENCY_FMT
            cell.font = TOTAL_FONT
            cell.fill = TOTAL_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="right")

    # Freeze header
    ws.freeze_panes = "A2"


def _write_summary_sheet(ws, income: list, expenses: list,
                         afm: str, date_from: str, date_to: str):
    """Write summary sheet with category breakdown, VAT, and result."""
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16

    row = 1

    # ── Section: Metadata ──
    _section_header(ws, row, "Στοιχεία Αναφοράς")
    row += 1
    ws.cell(row=row, column=1, value="ΑΦΜ:").font = LABEL_FONT
    ws.cell(row=row, column=2, value=afm)
    row += 1
    ws.cell(row=row, column=1, value="Περίοδος:").font = LABEL_FONT
    ws.cell(row=row, column=2, value=f"{date_from} — {date_to}")
    row += 1
    ws.cell(row=row, column=1, value="Ημ/νία δημιουργίας:").font = LABEL_FONT
    ws.cell(row=row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M"))
    row += 2

    # ── Section: Income by category ──
    _section_header(ws, row, "Έσοδα ανά Κατηγορία")
    row += 1
    for col_idx, title in enumerate(["Τύπος", "Περιγραφή", "Πλήθος", "Καθαρή", "ΦΠΑ", "Μικτή"], 1):
        cell = ws.cell(row=row, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    row += 1
    income_groups = _group_by_type(income)
    for inv_type, data in sorted(income_groups.items()):
        ws.cell(row=row, column=1, value=inv_type)
        ws.cell(row=row, column=2, value=INVOICE_TYPE_NAMES.get(inv_type, inv_type))
        ws.cell(row=row, column=3, value=data["count"])
        ws.cell(row=row, column=4, value=data["net"]).number_format = CURRENCY_FMT
        ws.cell(row=row, column=5, value=data["vat"]).number_format = CURRENCY_FMT
        ws.cell(row=row, column=6, value=data["gross"]).number_format = CURRENCY_FMT
        row += 1
    row += 1

    # ── Section: Expenses by category ──
    _section_header(ws, row, "Έξοδα ανά Κατηγορία")
    row += 1
    for col_idx, title in enumerate(["Τύπος", "Περιγραφή", "Πλήθος", "Καθαρή", "ΦΠΑ", "Μικτή"], 1):
        cell = ws.cell(row=row, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    row += 1
    expense_groups = _group_by_type(expenses)
    for inv_type, data in sorted(expense_groups.items()):
        ws.cell(row=row, column=1, value=inv_type)
        ws.cell(row=row, column=2, value=INVOICE_TYPE_NAMES.get(inv_type, inv_type))
        ws.cell(row=row, column=3, value=data["count"])
        ws.cell(row=row, column=4, value=data["net"]).number_format = CURRENCY_FMT
        ws.cell(row=row, column=5, value=data["vat"]).number_format = CURRENCY_FMT
        ws.cell(row=row, column=6, value=data["gross"]).number_format = CURRENCY_FMT
        row += 1
    row += 1

    # ── Section: VAT ──
    total_income_net = sum(d["net"] for d in income_groups.values())
    total_income_vat = sum(d["vat"] for d in income_groups.values())
    total_income_gross = sum(d["gross"] for d in income_groups.values())
    total_expense_net = sum(d["net"] for d in expense_groups.values())
    total_expense_vat = sum(d["vat"] for d in expense_groups.values())
    total_expense_gross = sum(d["gross"] for d in expense_groups.values())

    _section_header(ws, row, "ΦΠΑ")
    row += 1
    ws.cell(row=row, column=1, value="ΦΠΑ Εσόδων").font = LABEL_FONT
    ws.cell(row=row, column=2, value=total_income_vat).number_format = CURRENCY_FMT
    row += 1
    ws.cell(row=row, column=1, value="ΦΠΑ Εξόδων").font = LABEL_FONT
    ws.cell(row=row, column=2, value=total_expense_vat).number_format = CURRENCY_FMT
    row += 1
    vat_diff = total_income_vat - total_expense_vat
    ws.cell(row=row, column=1, value="Διαφορά ΦΠΑ").font = TOTAL_FONT
    cell = ws.cell(row=row, column=2, value=vat_diff)
    cell.number_format = CURRENCY_FMT
    cell.font = TOTAL_FONT
    cell.fill = TOTAL_FILL
    row += 2

    # ── Section: Result ──
    _section_header(ws, row, "Αποτέλεσμα")
    row += 1
    ws.cell(row=row, column=1, value="Σύνολο Εσόδων (καθαρά)").font = LABEL_FONT
    ws.cell(row=row, column=2, value=total_income_net).number_format = CURRENCY_FMT
    row += 1
    ws.cell(row=row, column=1, value="Σύνολο Εξόδων (καθαρά)").font = LABEL_FONT
    ws.cell(row=row, column=2, value=total_expense_net).number_format = CURRENCY_FMT
    row += 1
    profit = total_income_net - total_expense_net
    label = "Κέρδος" if profit >= 0 else "Ζημία"
    ws.cell(row=row, column=1, value=label).font = PROFIT_FONT if profit >= 0 else LOSS_FONT
    cell = ws.cell(row=row, column=2, value=profit)
    cell.number_format = CURRENCY_FMT
    cell.font = PROFIT_FONT if profit >= 0 else LOSS_FONT
    cell.fill = TOTAL_FILL
    row += 2

    # ── Section: Totals summary ──
    _section_header(ws, row, "Σύνολα")
    row += 1
    for col_idx, title in enumerate(["", "Καθαρή", "ΦΠΑ", "Μικτή"], 1):
        cell = ws.cell(row=row, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    row += 1
    ws.cell(row=row, column=1, value="Έσοδα").font = LABEL_FONT
    ws.cell(row=row, column=2, value=total_income_net).number_format = CURRENCY_FMT
    ws.cell(row=row, column=3, value=total_income_vat).number_format = CURRENCY_FMT
    ws.cell(row=row, column=4, value=total_income_gross).number_format = CURRENCY_FMT
    row += 1
    ws.cell(row=row, column=1, value="Έξοδα").font = LABEL_FONT
    ws.cell(row=row, column=2, value=total_expense_net).number_format = CURRENCY_FMT
    ws.cell(row=row, column=3, value=total_expense_vat).number_format = CURRENCY_FMT
    ws.cell(row=row, column=4, value=total_expense_gross).number_format = CURRENCY_FMT


def _section_header(ws, row: int, title: str):
    """Write a section header spanning columns A-F."""
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = SUMMARY_HEADER_FONT
    cell.fill = SUMMARY_HEADER_FILL
    for col in range(2, 7):
        ws.cell(row=row, column=col).fill = SUMMARY_HEADER_FILL


def _group_by_type(invoices: list) -> dict:
    """Group invoices by type, summing net/vat/gross."""
    groups = defaultdict(lambda: {"count": 0, "net": 0.0, "vat": 0.0, "gross": 0.0})
    for inv in invoices:
        inv_type = inv.get("invoiceType", "unknown")
        groups[inv_type]["count"] += 1
        groups[inv_type]["net"] += _to_float(inv.get("totalNetValue", 0))
        groups[inv_type]["vat"] += _to_float(inv.get("totalVatAmount", 0))
        groups[inv_type]["gross"] += _to_float(inv.get("totalGrossValue", 0))
    return dict(groups)


def _resolve_names(invoices: list, direction: str):
    """Resolve counterpart/issuer names from AFM via GSIS SOAP + cache."""
    try:
        from agent import _validate_afm, _afm_cache_get, _afm_cache_put, _soap_lookup_afm
    except Exception:
        return

    for inv in invoices:
        if direction == "sent":
            afm = inv.get("counterpart_vat", "")
            name = inv.get("counterpart_name", "")
        else:
            afm = inv.get("issuer_vat", "")
            name = inv.get("issuer_name", "")

        if name or not afm:
            continue

        afm_clean = afm.strip().replace("EL", "").replace("el", "")
        if not _validate_afm(afm_clean):
            continue

        cached = _afm_cache_get(afm_clean)
        if cached:
            resolved = cached.get("name", "")
        else:
            try:
                info = _soap_lookup_afm(afm_clean)
                _afm_cache_put(afm_clean, info)
                resolved = info.get("name", "")
            except Exception:
                continue

        if direction == "sent":
            inv["counterpart_name"] = resolved
        else:
            inv["issuer_name"] = resolved


def _group_by_supplier(expenses: list) -> dict:
    """Group expenses by supplier AFM, summing net/vat/gross."""
    groups = defaultdict(lambda: {"name": "", "count": 0, "net": 0.0, "vat": 0.0, "gross": 0.0, "invoices": []})
    for inv in expenses:
        afm = inv.get("issuer_vat", "unknown")
        name = inv.get("issuer_name", "")
        groups[afm]["count"] += 1
        if name and not groups[afm]["name"]:
            groups[afm]["name"] = name
        groups[afm]["net"] += _to_float(inv.get("totalNetValue", 0))
        groups[afm]["vat"] += _to_float(inv.get("totalVatAmount", 0))
        groups[afm]["gross"] += _to_float(inv.get("totalGrossValue", 0))
        groups[afm]["invoices"].append(inv)
    return dict(groups)


def _generate_supplier_xlsx(path: Path, expenses: list,
                            afm: str, date_from: str, date_to: str) -> Path:
    """Generate Excel with expenses grouped by supplier."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # Sheet 1: Σύνοψη ανά Προμηθευτή
    ws = wb.active
    ws.title = "Ανά Προμηθευτή"

    cols = [
        ("ΑΦΜ Προμηθευτή", 16),
        ("Επωνυμία", 36),
        ("Πλήθος", 10),
        ("Καθαρή Αξία", 16),
        ("ΦΠΑ", 14),
        ("Μικτή Αξία", 16),
    ]
    for i, (title, width) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
        cell = ws.cell(row=1, column=i, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    supplier_groups = _group_by_supplier(expenses)
    # Sort by gross descending
    sorted_suppliers = sorted(supplier_groups.items(), key=lambda x: x[1]["gross"], reverse=True)

    for row_idx, (supplier_afm, data) in enumerate(sorted_suppliers, 2):
        values = [supplier_afm, data["name"], data["count"], data["net"], data["vat"], data["gross"]]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            if col_idx >= 4:
                cell.number_format = CURRENCY_FMT
                cell.alignment = Alignment(horizontal="right")
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

    # Totals
    if sorted_suppliers:
        total_row = len(sorted_suppliers) + 2
        ws.cell(row=total_row, column=2, value="ΣΥΝΟΛΟ").font = TOTAL_FONT
        ws.cell(row=total_row, column=3, value=sum(d["count"] for d in supplier_groups.values())).font = TOTAL_FONT
        for col_idx in (4, 5, 6):
            col_letter = get_column_letter(col_idx)
            cell = ws.cell(
                row=total_row, column=col_idx,
                value=f"=SUM({col_letter}2:{col_letter}{total_row - 1})",
            )
            cell.number_format = CURRENCY_FMT
            cell.font = TOTAL_FONT
            cell.fill = TOTAL_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="right")

    ws.freeze_panes = "A2"

    # Sheet 2: Αναλυτικά Έξοδα
    ws_detail = wb.create_sheet("Αναλυτικά")
    _write_invoice_sheet(ws_detail, expenses, "expenses")

    # Sheet 3: Metadata
    ws_meta = wb.create_sheet("Στοιχεία")
    ws_meta.column_dimensions["A"].width = 28
    ws_meta.column_dimensions["B"].width = 30
    _section_header(ws_meta, 1, "Στοιχεία Αναφοράς")
    ws_meta.cell(row=2, column=1, value="ΑΦΜ:").font = LABEL_FONT
    ws_meta.cell(row=2, column=2, value=afm)
    ws_meta.cell(row=3, column=1, value="Περίοδος:").font = LABEL_FONT
    ws_meta.cell(row=3, column=2, value=f"{date_from} — {date_to}")
    ws_meta.cell(row=4, column=1, value="Ημ/νία δημιουργίας:").font = LABEL_FONT
    ws_meta.cell(row=4, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M"))
    ws_meta.cell(row=5, column=1, value="Προμηθευτές:").font = LABEL_FONT
    ws_meta.cell(row=5, column=2, value=str(len(supplier_groups)))
    ws_meta.cell(row=6, column=1, value="Σύνολο Παραστατικών:").font = LABEL_FONT
    ws_meta.cell(row=6, column=2, value=str(len(expenses)))

    wb.save(str(path))
    return path


def _to_float(val) -> float:
    """Safely convert value to float."""
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0
