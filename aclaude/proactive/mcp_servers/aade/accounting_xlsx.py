"""
AADE Accounting Excel Generator
================================
Generates an Excel workbook with 3 sheets:
  - Έσοδα (Income)
  - Έξοδα (Expenses)
  - Σύνοψη (Summary)
"""
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Invoice type names ──

INVOICE_TYPE_NAMES = {
    "1.1": "Τιμολόγιο Πώλησης",
    "1.2": "Τιμολόγιο Πώλησης / Ενδ.",
    "1.3": "Τιμολόγιο Πώλησης / Τρίτων",
    "1.6": "Τιμολόγιο Αυτοπαράδοσης",
    "2.1": "ΤΠΥ",
    "2.2": "ΤΠΥ / Ενδ.",
    "2.3": "ΤΠΥ / Τρίτων",
    "2.4": "Συμβόλαιο - Έσοδο",
    "3.1": "Τίτλος Κτήσης",
    "5.1": "Πιστωτικό (συσχ.)",
    "5.2": "Πιστωτικό (μη συσχ.)",
    "11.1": "ΑΛΠ",
    "11.2": "ΑΠΥ",
}

# ── Styles ──

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
SUMMARY_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUMMARY_HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=11)
PROFIT_FONT = Font(bold=True, size=12, color="006100")
LOSS_FONT = Font(bold=True, size=12, color="9C0006")
LABEL_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
CURRENCY_FMT = '#,##0.00 €'


def generate_xlsx(path: str | Path, income: list, expenses: list,
                  afm: str, date_from: str, date_to: str) -> Path:
    """Generate accounting Excel workbook.

    Args:
        path: Output file path
        income: List of income invoice dicts
        expenses: List of expense invoice dicts
        afm: Business AFM
        date_from: Period start (YYYY-MM-DD)
        date_to: Period end (YYYY-MM-DD)

    Returns:
        Path to the generated file
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # Sheet 1: Έσοδα
    ws_income = wb.active
    ws_income.title = "Έσοδα"
    _write_invoice_sheet(ws_income, income, "income")

    # Sheet 2: Έξοδα
    ws_expenses = wb.create_sheet("Έξοδα")
    _write_invoice_sheet(ws_expenses, expenses, "expenses")

    # Sheet 3: Σύνοψη
    ws_summary = wb.create_sheet("Σύνοψη")
    _write_summary_sheet(ws_summary, income, expenses, afm, date_from, date_to)

    wb.save(str(path))
    return path


# ── Column definitions ──

INCOME_COLUMNS = [
    ("Ημ/νία", 12),
    ("Σειρά/ΑΑ", 14),
    ("Τύπος", 8),
    ("Περιγραφή", 28),
    ("ΑΦΜ Πελάτη", 14),
    ("Καθαρή Αξία", 15),
    ("ΦΠΑ", 12),
    ("Μικτή Αξία", 15),
    ("MARK", 14),
]

EXPENSE_COLUMNS = [
    ("Ημ/νία", 12),
    ("Σειρά/ΑΑ", 14),
    ("Τύπος", 8),
    ("Περιγραφή", 28),
    ("ΑΦΜ Προμηθευτή", 16),
    ("Επωνυμία Προμηθευτή", 32),
    ("Καθαρή Αξία", 15),
    ("ΦΠΑ", 12),
    ("Μικτή Αξία", 15),
    ("MARK", 14),
]


def _write_invoice_sheet(ws, invoices: list, direction: str):
    """Write income or expenses invoice sheet."""
    columns = INCOME_COLUMNS if direction == "income" else EXPENSE_COLUMNS

    # Column widths
    for i, (_, width) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Header row
    for col_idx, (title, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Data rows
    for row_idx, inv in enumerate(invoices, 2):
        net = _to_float(inv.get("totalNetValue", 0))
        vat = _to_float(inv.get("totalVatAmount", 0))
        gross = _to_float(inv.get("totalGrossValue", 0))
        inv_type = inv.get("invoiceType", "")
        type_name = INVOICE_TYPE_NAMES.get(inv_type, inv.get("type_name", inv_type))
        series_aa = f"{inv.get('series', '')}/{inv.get('aa', '')}"

        if direction == "expenses":
            values = [
                inv.get("issueDate", ""),
                series_aa,
                inv_type,
                type_name,
                inv.get("issuer_vat", ""),
                inv.get("issuer_name", ""),
                net,
                vat,
                gross,
                inv.get("mark", ""),
            ]
        else:
            values = [
                inv.get("issueDate", ""),
                series_aa,
                inv_type,
                type_name,
                inv.get("counterpart_vat", inv.get("issuer_vat", "")),
                net,
                vat,
                gross,
                inv.get("mark", ""),
            ]

        # Determine which columns are currency (depends on direction)
        num_cols = len(columns)
        # Currency columns are always the 3rd, 2nd, and 1st before MARK
        currency_cols = (num_cols - 3, num_cols - 2, num_cols - 1)  # net, vat, gross

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            # Currency formatting for value columns
            if col_idx in currency_cols:
                cell.number_format = CURRENCY_FMT
                cell.alignment = Alignment(horizontal="right")
            # Alternating row colors
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

    # Totals row
    if invoices:
        total_row = len(invoices) + 2
        num_cols = len(columns)
        ws.cell(row=total_row, column=4, value="ΣΥΝΟΛΟ").font = TOTAL_FONT

        # Currency columns: net, vat, gross (3rd, 2nd, 1st before MARK)
        for col_idx in (num_cols - 3, num_cols - 2, num_cols - 1):
            col_letter = get_column_letter(col_idx)
            cell = ws.cell(
                row=total_row, column=col_idx,
                value=f"=SUM({col_letter}2:{col_letter}{total_row - 1})",
            )
            cell.number_format = CURRENCY_FMT
            cell.font = TOTAL_FONT
            cell.fill = TOTAL_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="right")

    # Freeze header
    ws.freeze_panes = "A2"


def _write_summary_sheet(ws, income: list, expenses: list,
                         afm: str, date_from: str, date_to: str):
    """Write summary sheet with category breakdown, VAT, and result."""
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16

    row = 1

    # ── Section: Metadata ──
    _section_header(ws, row, "Στοιχεία Αναφοράς")
    row += 1
    ws.cell(row=row, column=1, value="ΑΦΜ:").font = LABEL_FONT
    ws.cell(row=row, column=2, value=afm)
    row += 1
    ws.cell(row=row, column=1, value="Περίοδος:").font = LABEL_FONT
    ws.cell(row=row, column=2, value=f"{date_from} — {date_to}")
    row += 1
    ws.cell(row=row, column=1, value="Ημ/νία δημιουργίας:").font = LABEL_FONT
    ws.cell(row=row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M"))
    row += 2

    # ── Section: Income by category ──
    _section_header(ws, row, "Έσοδα ανά Κατηγορία")
    row += 1
    for col_idx, title in enumerate(["Τύπος", "Περιγραφή", "Πλήθος", "Καθαρή", "ΦΠΑ", "Μικτή"], 1):
        cell = ws.cell(row=row, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    row += 1
    income_groups = _group_by_type(income)
    for inv_type, data in sorted(income_groups.items()):
        ws.cell(row=row, column=1, value=inv_type)
        ws.cell(row=row, column=2, value=INVOICE_TYPE_NAMES.get(inv_type, inv_type))
        ws.cell(row=row, column=3, value=data["count"])
        ws.cell(row=row, column=4, value=data["net"]).number_format = CURRENCY_FMT
        ws.cell(row=row, column=5, value=data["vat"]).number_format = CURRENCY_FMT
        ws.cell(row=row, column=6, value=data["gross"]).number_format = CURRENCY_FMT
        row += 1
    row += 1

    # ── Section: Expenses by category ──
    _section_header(ws, row, "Έξοδα ανά Κατηγορία")
    row += 1
    for col_idx, title in enumerate(["Τύπος", "Περιγραφή", "Πλήθος", "Καθαρή", "ΦΠΑ", "Μικτή"], 1):
        cell = ws.cell(row=row, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    row += 1
    expense_groups = _group_by_type(expenses)
    for inv_type, data in sorted(expense_groups.items()):
        ws.cell(row=row, column=1, value=inv_type)
        ws.cell(row=row, column=2, value=INVOICE_TYPE_NAMES.get(inv_type, inv_type))
        ws.cell(row=row, column=3, value=data["count"])
        ws.cell(row=row, column=4, value=data["net"]).number_format = CURRENCY_FMT
        ws.cell(row=row, column=5, value=data["vat"]).number_format = CURRENCY_FMT
        ws.cell(row=row, column=6, value=data["gross"]).number_format = CURRENCY_FMT
        row += 1
    row += 1

    # ── Section: VAT ──
    total_income_net = sum(d["net"] for d in income_groups.values())
    total_income_vat = sum(d["vat"] for d in income_groups.values())
    total_income_gross = sum(d["gross"] for d in income_groups.values())
    total_expense_net = sum(d["net"] for d in expense_groups.values())
    total_expense_vat = sum(d["vat"] for d in expense_groups.values())
    total_expense_gross = sum(d["gross"] for d in expense_groups.values())

    _section_header(ws, row, "ΦΠΑ")
    row += 1
    ws.cell(row=row, column=1, value="ΦΠΑ Εσόδων").font = LABEL_FONT
    ws.cell(row=row, column=2, value=total_income_vat).number_format = CURRENCY_FMT
    row += 1
    ws.cell(row=row, column=1, value="ΦΠΑ Εξόδων").font = LABEL_FONT
    ws.cell(row=row, column=2, value=total_expense_vat).number_format = CURRENCY_FMT
    row += 1
    vat_diff = total_income_vat - total_expense_vat
    ws.cell(row=row, column=1, value="Διαφορά ΦΠΑ").font = TOTAL_FONT
    cell = ws.cell(row=row, column=2, value=vat_diff)
    cell.number_format = CURRENCY_FMT
    cell.font = TOTAL_FONT
    cell.fill = TOTAL_FILL
    row += 2

    # ── Section: Result ──
    _section_header(ws, row, "Αποτέλεσμα")
    row += 1
    ws.cell(row=row, column=1, value="Σύνολο Εσόδων (καθαρά)").font = LABEL_FONT
    ws.cell(row=row, column=2, value=total_income_net).number_format = CURRENCY_FMT
    row += 1
    ws.cell(row=row, column=1, value="Σύνολο Εξόδων (καθαρά)").font = LABEL_FONT
    ws.cell(row=row, column=2, value=total_expense_net).number_format = CURRENCY_FMT
    row += 1
    profit = total_income_net - total_expense_net
    label = "Κέρδος" if profit >= 0 else "Ζημία"
    ws.cell(row=row, column=1, value=label).font = PROFIT_FONT if profit >= 0 else LOSS_FONT
    cell = ws.cell(row=row, column=2, value=profit)
    cell.number_format = CURRENCY_FMT
    cell.font = PROFIT_FONT if profit >= 0 else LOSS_FONT
    cell.fill = TOTAL_FILL
    row += 2

    # ── Section: Totals summary ──
    _section_header(ws, row, "Σύνολα")
    row += 1
    for col_idx, title in enumerate(["", "Καθαρή", "ΦΠΑ", "Μικτή"], 1):
        cell = ws.cell(row=row, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    row += 1
    ws.cell(row=row, column=1, value="Έσοδα").font = LABEL_FONT
    ws.cell(row=row, column=2, value=total_income_net).number_format = CURRENCY_FMT
    ws.cell(row=row, column=3, value=total_income_vat).number_format = CURRENCY_FMT
    ws.cell(row=row, column=4, value=total_income_gross).number_format = CURRENCY_FMT
    row += 1
    ws.cell(row=row, column=1, value="Έξοδα").font = LABEL_FONT
    ws.cell(row=row, column=2, value=total_expense_net).number_format = CURRENCY_FMT
    ws.cell(row=row, column=3, value=total_expense_vat).number_format = CURRENCY_FMT
    ws.cell(row=row, column=4, value=total_expense_gross).number_format = CURRENCY_FMT


def _section_header(ws, row: int, title: str):
    """Write a section header spanning columns A-F."""
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = SUMMARY_HEADER_FONT
    cell.fill = SUMMARY_HEADER_FILL
    for col in range(2, 7):
        ws.cell(row=row, column=col).fill = SUMMARY_HEADER_FILL


def _group_by_type(invoices: list) -> dict:
    """Group invoices by type, summing net/vat/gross."""
    groups = defaultdict(lambda: {"count": 0, "net": 0.0, "vat": 0.0, "gross": 0.0})
    for inv in invoices:
        inv_type = inv.get("invoiceType", "unknown")
        groups[inv_type]["count"] += 1
        groups[inv_type]["net"] += _to_float(inv.get("totalNetValue", 0))
        groups[inv_type]["vat"] += _to_float(inv.get("totalVatAmount", 0))
        groups[inv_type]["gross"] += _to_float(inv.get("totalGrossValue", 0))
    return dict(groups)


def _to_float(val) -> float:
    """Safely convert value to float."""
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0
